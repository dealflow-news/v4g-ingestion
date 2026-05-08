#!/usr/bin/env python3
"""
Lane B β0 — Export financials for one party from Supabase to Excel.

Usage:
    python -m src.cli.export_financials_xlsx --party-id <uuid> [--out PATH]

Reads `public.vw_target_financials` (analytical layer with YoY deltas + ratios).
Read-only — no DB writes. Safe to run locally with .env service_role key.

Output: 3-sheet .xlsx
    • Pivot       — metrics × fiscal years (analyst-friendly)
    • Raw         — one row per fiscal year, all columns (filterable)
    • Provenance  — filing_date, source_code, confidence per period (audit trail)

Doctrine: this CLI is the "Export out of Supabase" leg of the three-act
flow. Analysis lives in DB views; CLI only queries, formats, and delivers.
"""
from __future__ import annotations

import logging
from pathlib import Path
from uuid import UUID

import click
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.persistence.supabase import admin_client

# Auto-load .env from repo root so admin_client() finds SUPABASE_URL etc.
# override=False keeps any pre-set shell env vars authoritative.
load_dotenv(Path(__file__).resolve().parents[2] / ".env", override=False)

log = logging.getLogger(__name__)
logging.basicConfig(
    level="INFO",
    format="%(asctime)s  %(levelname)-5s  %(message)s",
)


# ─── Layout config ──────────────────────────────────────────────────────
# (column in vw_target_financials, display label, openpyxl number format)
# Rows starting with __sep__ render as a visual separator band.
PIVOT_METRICS: list[tuple[str, str | None, str | None]] = [
    ("revenue_eur_m",                "Revenue (EUR M)",            "#,##0.00"),
    ("revenue_yoy_delta_eur_m",      "  YoY Δ revenue (EUR M)",    "+#,##0.00;-#,##0.00"),
    ("ebitda_eur_m",                 "EBITDA (EUR M)",             "#,##0.00"),
    ("ebitda_yoy_delta_eur_m",       "  YoY Δ EBITDA (EUR M)",     "+#,##0.00;-#,##0.00"),
    ("ebitda_margin_pct",            "  EBITDA margin (%)",        "0.0"),
    ("ebit_eur_m",                   "EBIT (EUR M)",               "#,##0.00"),
    ("net_income_eur_m",             "Net income (EUR M)",         "#,##0.00"),
    ("__sep1__",                     None,                          None),
    ("total_assets_eur_m",           "Total assets (EUR M)",       "#,##0.00"),
    ("total_equity_eur_m",           "Total equity (EUR M)",       "#,##0.00"),
    ("cash_eur_m",                   "Cash (EUR M)",               "#,##0.00"),
    ("total_debt_eur_m",             "Total debt (EUR M)",         "#,##0.00"),
    ("net_debt_eur_m",               "Net debt (EUR M)",           "#,##0.00"),
    ("net_leverage_x",               "  Net leverage (x)",         "0.0"),
    ("working_capital_eur_m",        "Working capital (EUR M)",    "#,##0.00"),
    ("__sep2__",                     None,                          None),
    ("employees",                    "Employees (FTE)",            "#,##0"),
    ("revenue_per_employee_eur_k",   "  Revenue / FTE (EUR K)",    "#,##0"),
    ("ebitda_per_employee_eur_k",    "  EBITDA / FTE (EUR K)",     "#,##0"),
]

PROVENANCE_COLS = [
    "period_end", "period_label", "period_type",
    "source_code", "confidence",
    "nbb_model_type", "nbb_filing_date",
]

# V4G amber + dark navy palette
_HEADER_FILL = PatternFill("solid", fgColor="E8A020")
_HEADER_FONT = Font(bold=True, color="0F1520")
_TITLE_FONT  = Font(bold=True, size=14, color="E8A020")
_MUTED_FONT  = Font(color="64748B")
_SEP_FILL    = PatternFill("solid", fgColor="1E2D45")


# ─── Helpers ────────────────────────────────────────────────────────────
def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, min_w: int = 10, max_w: int = 28) -> None:
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)


# ─── Data fetch ─────────────────────────────────────────────────────────
def fetch_financials(party_id: str) -> tuple[list[dict], dict | None]:
    """Returns (rows from vw_target_financials, party metadata enriched w/ KBO).

    KBO comes from `party_identifiers` (id_type='KBO') — canonical lookup
    pattern matching `runner.py._resolve_kbo`. `party_registry` itself has
    no `kbo_nr` column.
    """
    client = admin_client()

    # Base party metadata (no kbo_nr column on this table)
    party_resp = (
        client.table("party_registry")
        .select("party_id, legal_name, country_iso2, party_type")
        .eq("party_id", party_id)
        .limit(1)
        .execute()
    )
    party_meta = party_resp.data[0] if party_resp.data else None

    # Enrich with KBO via party_identifiers
    if party_meta:
        kbo_resp = (
            client.table("party_identifiers")
            .select("id_value")
            .eq("party_id", party_id)
            .eq("id_type", "KBO")
            .limit(1)
            .execute()
        )
        party_meta["kbo_nr"] = kbo_resp.data[0]["id_value"] if kbo_resp.data else None

    fin_resp = (
        client.table("vw_target_financials")
        .select("*")
        .eq("party_id", party_id)
        .order("period_end", desc=True)
        .execute()
    )
    return fin_resp.data or [], party_meta


# ─── Sheet builders ─────────────────────────────────────────────────────
def build_pivot_sheet(ws, rows: list[dict], party_meta: dict | None) -> None:
    """Metrics as rows, fiscal years as columns. Chronological L→R."""
    ws.title = "Pivot"

    name = (party_meta or {}).get("legal_name", "(unknown party)")
    kbo  = (party_meta or {}).get("kbo_nr") or ""
    ws["A1"] = f"Financial pivot · {name}"
    ws["A1"].font = _TITLE_FONT
    sub = f"KBO: {kbo} · " if kbo else ""
    ws["A2"] = f"{sub}Source: vw_target_financials  ·  rows: {len(rows)}"
    ws["A2"].font = _MUTED_FONT

    chrono = sorted(rows, key=lambda r: r.get("period_end") or "9999-99-99")
    years = [r.get("period_label") or (r.get("period_end") or "")[:4] for r in chrono]

    ws.cell(row=4, column=1, value="Metric")
    for i, year in enumerate(years, start=2):
        ws.cell(row=4, column=i, value=year)
    _style_header_row(ws, 4, len(years) + 1)

    row_idx = 5
    for col, label, fmt in PIVOT_METRICS:
        if col.startswith("__sep"):
            for j in range(1, len(years) + 2):
                ws.cell(row=row_idx, column=j).fill = _SEP_FILL
            row_idx += 1
            continue
        ws.cell(row=row_idx, column=1, value=label)
        for j, r in enumerate(chrono, start=2):
            v = r.get(col)
            cell = ws.cell(row=row_idx, column=j, value=v)
            if fmt and v is not None:
                cell.number_format = fmt
        row_idx += 1

    ws.freeze_panes = "B5"
    _autosize(ws, min_w=14, max_w=22)


def build_raw_sheet(ws, rows: list[dict]) -> None:
    """Flat one-row-per-year dump."""
    ws.title = "Raw"
    if not rows:
        ws["A1"] = "(no rows)"
        return

    cols = list(rows[0].keys())
    for i, col in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=col)
    _style_header_row(ws, 1, len(cols))

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))

    ws.freeze_panes = "B2"
    _autosize(ws, min_w=10, max_w=24)


def build_provenance_sheet(ws, rows: list[dict]) -> None:
    """Audit trail per fiscal year."""
    ws.title = "Provenance"

    ws["A1"] = "Per-period provenance"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "filing_date / source / confidence per fiscal year"
    ws["A2"].font = _MUTED_FONT

    for i, col in enumerate(PROVENANCE_COLS, start=1):
        ws.cell(row=4, column=i, value=col)
    _style_header_row(ws, 4, len(PROVENANCE_COLS))

    chrono = sorted(rows, key=lambda r: r.get("period_end") or "0000-00-00", reverse=True)
    for r_idx, row in enumerate(chrono, start=5):
        for c_idx, col in enumerate(PROVENANCE_COLS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))

    ws.freeze_panes = "A5"
    _autosize(ws, min_w=14, max_w=22)


# ─── CLI ────────────────────────────────────────────────────────────────
@click.command()
@click.option("--party-id", required=True, help="UUID of the party to export")
@click.option(
    "--out",
    default=None,
    type=click.Path(path_type=Path),
    help="Output path (default: ./financials_<party-short>.xlsx)",
)
def export(party_id: str, out: Path | None) -> None:
    """Export financials for one party to a 3-sheet Excel workbook."""
    try:
        UUID(party_id)
    except ValueError as err:
        raise click.BadParameter(f"Invalid UUID: {party_id}") from err

    log.info("fetching financials party_id=%s", party_id)
    rows, party_meta = fetch_financials(party_id)

    if not rows:
        click.echo(f"\nNo fact_financials rows for party {party_id}.")
        click.echo(
            "Either: party has no SRC_NBB / SRC_PB data yet, or the UUID is wrong."
        )
        raise click.Abort()

    log.info(
        "found %d rows · party=%s · KBO=%s",
        len(rows),
        (party_meta or {}).get("legal_name", "?"),
        (party_meta or {}).get("kbo_nr", "?"),
    )

    wb = Workbook()
    build_pivot_sheet(wb.active, rows, party_meta)
    build_raw_sheet(wb.create_sheet(), rows)
    build_provenance_sheet(wb.create_sheet(), rows)

    if out is None:
        out = Path(f"financials_{party_id[:8]}.xlsx")
    wb.save(out)

    log.info("wrote %s · %d years · 3 sheets", out.resolve(), len(rows))
    click.echo(f"\n✓ {out}  ({len(rows)} years, 3 sheets)")


if __name__ == "__main__":
    export()
