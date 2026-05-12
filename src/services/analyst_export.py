"""PCMN-detailed M&A analyst export.

Generates an .xlsx file with three sheets:

* ``Info`` -- party metadata + report parameters
* ``Yearly_Review`` -- PCMN-detailed financial data laid out as
  (sections of codes) x (most-recent N reporting periods). ~80 rows x N columns.
* ``Filings`` -- raw filings table for transparency

Replaces the previous aggregated analyst format (P&L / Balance Sheet / KPIs
sheets). KPIs with sector benchmarks become the scope of the "simple" route.

Mirrors the ``ExcelExporter`` protocol: ``fetch() -> build() -> bytes``.

Data sources:

* ``party_registry``       -- party display info
* ``party_identifiers``    -- primary KBO
* ``fact_filings``         -- period list (filtered: superseded_by IS NULL)
* ``dim_pcmn_codes``       -- row template (sections, labels, display_order)
* ``fact_financials_lines``-- amount per (filing_id, pcmn_code), pivoted in Python
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)


# Canonical section ordering (locked by W8-EXT-002 doctrine).
# NOTES is excluded from analyst layout (not an analyst section).
SECTION_HEADERS: dict[str, str] = {
    "BS_A":    "BALANCE SHEET - ASSETS",
    "BS_L":    "BALANCE SHEET - LIABILITIES & EQUITY",
    "IS":      "INCOME STATEMENT",
    "IS_APPR": "PROFIT APPROPRIATION",
    "WORKERS": "WORKFORCE",
}
SECTION_ORDER: list[str] = ["BS_A", "BS_L", "IS", "IS_APPR", "WORKERS"]


# Styling
TITLE_FONT    = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SECTION_FONT  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FONT   = Font(name="Calibri", size=10, bold=True, color="1F4E79")
HIGH_FONT     = Font(name="Calibri", size=10, bold=True)
MED_FONT      = Font(name="Calibri", size=10)
LOW_FONT      = Font(name="Calibri", size=10, color="888888")
SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True)
WARNING_FONT  = Font(name="Calibri", size=9, italic=True, color="B7472A")
FOOTER_FONT   = Font(name="Calibri", size=8, italic=True, color="888888")
META_FONT     = Font(name="Calibri", size=10, bold=True, color="1F4E79")

SECTION_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL   = PatternFill("solid", fgColor="DEEBF7")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")

AMOUNT_FORMAT = '#,##0.0;(#,##0.0);"-"'   # EUR thousands, one decimal
COUNT_FORMAT  = '#,##0;(#,##0);"-"'       # integer counts (FTE, hours, etc.)

# Subsections from MAR_Dictionary that represent counts rather than EUR amounts.
# Used to decide between AMOUNT_FORMAT (divide by 1000) and COUNT_FORMAT (raw value).
COUNT_SUBSECTIONS = frozenset({"Count", "Headcount"})


@dataclass
class _PartyInfo:
    party_id: UUID
    display_name: str
    legal_name: str | None
    country_iso2: str | None
    party_type: str | None
    primary_kbo: str | None


@dataclass
class _Filing:
    filing_id: str
    filing_reference: str
    period_start: str  # ISO date
    period_end: str    # ISO date
    period_months: int
    period_flag: str
    period_label: str
    nbb_model_type: str | None
    deposit_date: str | None
    source_code: str
    enterprise_name: str | None
    currency: str
    derived_year: int  # extracted from period_end for column-header display


@dataclass
class _CodeRow:
    pcmn_code: str
    section: str
    label_en: str
    subsection: str | None
    v4g_priority: str
    display_order: int


@dataclass
class _Data:
    party: _PartyInfo
    filings: list[_Filing] = field(default_factory=list)
    codes: list[_CodeRow] = field(default_factory=list)
    # Pivoted amounts: amounts[pcmn_code][filing_id] = amount_eur
    amounts: dict[str, dict[str, float]] = field(default_factory=dict)


class AnalystExportError(Exception):
    """Raised when the analyst export cannot be produced."""


class AnalystExporter:
    """PCMN-detailed M&A analyst export. Three sheets: Info / Yearly_Review / Filings."""

    def __init__(self, client: Any, party_id: UUID | str, year_count: int = 10) -> None:
        self.client = client
        self.party_id = UUID(str(party_id)) if not isinstance(party_id, UUID) else party_id
        self.year_count = max(1, year_count)
        self._data: _Data | None = None

    # ----------------------------------------------------------- public API

    def fetch(self) -> AnalystExporter:
        """Pull party, filings, codes, and pivot amounts from Supabase."""
        party = self._fetch_party()
        filings = self._fetch_filings()
        if not filings:
            raise AnalystExportError(
                f"No filings found for party_id={self.party_id}"
            )
        codes = self._fetch_codes()
        amounts = self._fetch_amounts(
            filing_ids=[f.filing_id for f in filings],
            pcmn_codes=[c.pcmn_code for c in codes],
        )

        # Pivot: amounts[pcmn_code][filing_id] = amount_eur
        pivot: dict[str, dict[str, float]] = {}
        for row in amounts:
            code = row.get("pcmn_code")
            fid = row.get("filing_id")
            amt = row.get("amount_eur")
            if code is None or fid is None or amt is None:
                continue
            try:
                pivot.setdefault(code, {})[str(fid)] = float(amt)
            except (TypeError, ValueError):
                continue

        self._data = _Data(party=party, filings=filings, codes=codes, amounts=pivot)
        log.info(
            "analyst_export.fetch: party=%s filings=%d codes=%d amount_cells=%d",
            self.party_id, len(filings), len(codes),
            sum(len(v) for v in pivot.values()),
        )
        return self

    def build(self) -> bytes:
        """Render the workbook to bytes. Must call ``fetch()`` first."""
        if self._data is None:
            raise AnalystExportError("Call fetch() before build().")
        wb = Workbook()
        # Remove the default blank sheet
        default = wb.active
        if default is not None:
            wb.remove(default)
        self._build_info_sheet(wb)
        self._build_yearly_review_sheet(wb)
        self._build_filings_sheet(wb)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def suggest_filename(self) -> str:
        """Return a download-safe filename for the export."""
        if self._data is None:
            return f"V4G_analyst_{self.party_id}.xlsx"
        safe_name = "".join(
            c if c.isalnum() or c in "-_" else "_"
            for c in (self._data.party.display_name or "unknown")
        )[:60]
        kbo = self._data.party.primary_kbo or "NOKBO"
        return f"V4G_{safe_name}_{kbo}_analyst.xlsx"

    # -------------------------------------------------------- data fetching

    def _fetch_party(self) -> _PartyInfo:
        resp = (
            self.client.table("party_registry")
            .select("party_id, display_name, legal_name, country_iso2, party_type")
            .eq("party_id", str(self.party_id))
            .limit(1)
            .execute()
        )
        if not resp.data:
            raise AnalystExportError(
                f"Party {self.party_id} not found in party_registry"
            )
        row = resp.data[0]
        # KBO lookup -- match ExcelExporter convention: party_id + id_type only,
        # no is_primary filter (not all KBO rows have it set to TRUE).
        kbo_resp = (
            self.client.table("party_identifiers")
            .select("id_value")
            .eq("party_id", str(self.party_id))
            .eq("id_type", "KBO")
            .limit(1)
            .execute()
        )
        primary_kbo = kbo_resp.data[0]["id_value"] if kbo_resp.data else None
        display_name = (
            row.get("display_name") or row.get("legal_name") or str(self.party_id)
        )
        return _PartyInfo(
            party_id=self.party_id,
            display_name=display_name,
            legal_name=row.get("legal_name"),
            country_iso2=row.get("country_iso2"),
            party_type=row.get("party_type"),
            primary_kbo=primary_kbo,
        )

    def _fetch_filings(self) -> list[_Filing]:
        """Fetch filings for the party, drop superseded, dedupe by period_end, take year_count most recent.

        Uses ``period_*`` column naming (the actual schema). ``superseded_by`` IS
        NULL filter is applied Python-side to avoid supabase-py null-filter quirks.
        Dedup by ``period_end`` covers the edge case of two filings within one
        calendar year (e.g. fiscal-year boundary switch).
        """
        resp = (
            self.client.table("fact_filings")
            .select(
                "filing_id, filing_reference, period_start, period_end, "
                "period_months, period_flag, period_label, nbb_model_type, "
                "deposit_date, source_code, enterprise_name, currency, superseded_by"
            )
            .eq("party_id", str(self.party_id))
            .order("period_end", desc=True)
            .order("deposit_date", desc=True)
            .execute()
        )
        rows = resp.data or []

        # Python-side filter: drop superseded amendments
        rows = [r for r in rows if r.get("superseded_by") is None]

        seen_ends: set[str] = set()
        deduped: list[_Filing] = []
        for r in rows:
            end = r.get("period_end")
            if end is None or end in seen_ends:
                continue
            seen_ends.add(end)
            try:
                derived_year = int(str(end)[:4])
            except ValueError:
                derived_year = 0
            deduped.append(_Filing(
                filing_id=str(r["filing_id"]),
                filing_reference=str(r.get("filing_reference") or ""),
                period_start=str(r.get("period_start") or ""),
                period_end=str(end),
                period_months=int(r.get("period_months") or 0),
                period_flag=str(r.get("period_flag") or ""),
                period_label=str(r.get("period_label") or ""),
                nbb_model_type=r.get("nbb_model_type"),
                deposit_date=r.get("deposit_date"),
                source_code=str(r.get("source_code") or ""),
                enterprise_name=r.get("enterprise_name"),
                currency=str(r.get("currency") or "EUR"),
                derived_year=derived_year,
            ))
        return deduped[: self.year_count]

    def _fetch_codes(self) -> list[_CodeRow]:
        """Fetch dim_pcmn_codes rows in analyst order (filter to MAR codes only)."""
        resp = (
            self.client.table("dim_pcmn_codes")
            .select(
                "pcmn_code, section, label_en, subsection, v4g_priority, display_order"
            )
            .execute()
        )
        rows = resp.data or []
        codes: list[_CodeRow] = []
        for r in rows:
            # Python-side filter -- MAR_Dictionary codes only (label_en NOT NULL)
            # and only sections in our analyst layout.
            if not r.get("label_en"):
                continue
            if r.get("section") not in SECTION_ORDER:
                continue
            codes.append(_CodeRow(
                pcmn_code=str(r["pcmn_code"]),
                section=str(r["section"]),
                label_en=str(r.get("label_en") or ""),
                subsection=r.get("subsection"),
                v4g_priority=str(r.get("v4g_priority") or "MEDIUM"),
                display_order=int(r.get("display_order") or 0),
            ))
        sec_rank = {s: i for i, s in enumerate(SECTION_ORDER)}
        codes.sort(key=lambda c: (sec_rank[c.section], c.display_order))
        return codes

    def _fetch_amounts(
        self, filing_ids: list[str], pcmn_codes: list[str],
    ) -> list[dict]:
        """Fetch fact_financials_lines for the selected filings, then filter
        to analyst pcmn_codes Python-side.

        BUG HISTORY (v1, fixed in v2):
        v1 attempted to pre-filter codes server-side via ``.in_("pcmn_code", ...)``
        in the same query. With ~80 codes the resulting URL was apparently long
        enough that PostgREST silently dropped the filter (or supabase-py
        truncated it), so every line of every filing came back. For m02-f
        filings that's ~280 lines each = ~1400 rows for a 5-year export,
        which hit PostgREST's default db-max-rows=1000 cap. The pagination
        loop's second .range() request did fire but the first page already
        truncated mid-filing, producing the classic symptom: one filing
        (whichever sorted last in filing_id alphabetical order) had its
        higher-alphabet pcmn_codes silently dropped -- all P&L (60s, 70s)
        and workforce (9000s) codes blank in the analyst output.

        v2 (this version): one server query per page, NO pcmn_code filter
        server-side. Code filtering happens in Python after the rows are in
        memory. Same total bytes over the wire (~1400 rows for a typical
        5-filing m02-f export), but reliably across pages.

        Defensive measures:
        - ORDER BY (filing_id, pcmn_code) for deterministic pagination
        - Pagination loop with explicit page_size and safety stop
        - Python-side code filter via set lookup (O(1) per row)
        """
        if not filing_ids or not pcmn_codes:
            return []
        code_set = set(pcmn_codes)
        all_rows: list[dict] = []
        page_size = 1000
        offset = 0
        while True:
            resp = (
                self.client.table("fact_financials_lines")
                .select("filing_id, pcmn_code, amount_eur")
                .in_("filing_id", filing_ids)
                .order("filing_id")
                .order("pcmn_code")
                .range(offset, offset + page_size - 1)
                .execute()
            )
            rows = resp.data or []
            all_rows.extend(rows)
            if len(rows) < page_size:
                break
            offset += page_size
            if offset > 200_000:  # safety: 200k lines = ~700 filings worth
                log.warning(
                    "analyst_export._fetch_amounts: stopped at %d rows; "
                    "filing_ids=%d codes=%d",
                    offset, len(filing_ids), len(pcmn_codes),
                )
                break
        # Filter to analyst codes Python-side -- the only reliable place.
        matched = [r for r in all_rows if r.get("pcmn_code") in code_set]
        log.info(
            "analyst_export._fetch_amounts: %d filings x %d codes -> "
            "%d raw rows -> %d matched",
            len(filing_ids), len(pcmn_codes), len(all_rows), len(matched),
        )
        return matched

    # ------------------------------------------------------- sheet builders

    def _build_info_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Info")
        d = self._data
        assert d is not None  # noqa: S101 -- guarded by build()
        p = d.party

        ws["A1"] = "V4G Analyst Export"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:B1")

        rows = [
            ("Display Name",        p.display_name),
            ("Legal Name",          p.legal_name or "-"),
            ("Country",             p.country_iso2 or "-"),
            ("Party Type",          p.party_type or "-"),
            ("Primary KBO",         p.primary_kbo or "-"),
            ("Party ID",            str(p.party_id)),
            ("",                    ""),
            ("Years Requested",     str(self.year_count)),
            ("Filings Selected",    str(len(d.filings))),
            ("Earliest Period End", d.filings[-1].period_end if d.filings else "-"),
            ("Latest Period End",   d.filings[0].period_end if d.filings else "-"),
            ("Analyst Codes",       str(len(d.codes))),
            ("Amount Cells",        str(sum(len(v) for v in d.amounts.values()))),
            ("",                    ""),
            ("Generated At",        datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
            ("Generator",           "Golden Safe AnalystExporter v1"),
        ]
        for i, (label, value) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=label).font = META_FONT if label else Font()
            ws.cell(row=i, column=2, value=value)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 56

    def _build_yearly_review_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Yearly_Review")
        d = self._data
        assert d is not None  # noqa: S101

        filings = d.filings  # already period_end desc
        n_years = len(filings)
        last_col = 2 + n_years

        # Row 1: Title
        ws.cell(row=1, column=1, value=f"{d.party.display_name} - Yearly Review").font = TITLE_FONT
        if last_col > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

        # Row 2: KBO + generated timestamp
        meta = (
            f"KBO: {d.party.primary_kbo or '-'}  -  "
            f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}  -  "
            f"Amounts in EUR thousands; counts (FTE, hours) shown as integers"
        )
        ws.cell(row=2, column=1, value=meta).font = FOOTER_FONT
        if last_col > 1:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

        # Row 3: Period flag row (only non-12m periods get a marker)
        for j, f in enumerate(filings):
            cell = ws.cell(row=3, column=3 + j, value=self._period_flag(f.period_months))
            cell.font = WARNING_FONT
            cell.alignment = Alignment(horizontal="center")

        # Row 4: Header row (Code | Description | year1 | year2 | ...)
        ws.cell(row=4, column=1, value="Code").font = HEADER_FONT
        ws.cell(row=4, column=2, value="Description").font = HEADER_FONT
        ws.cell(row=4, column=1).fill = HEADER_FILL
        ws.cell(row=4, column=2).fill = HEADER_FILL
        for j, f in enumerate(filings):
            cell = ws.cell(row=4, column=3 + j, value=f.derived_year)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Body
        row = 5
        last_section: str | None = None
        for code in d.codes:
            if code.section != last_section:
                self._write_section_header(ws, row, last_col, code.section)
                last_section = code.section
                row += 1
            self._write_code_row(ws, row, code, filings, d.amounts)
            row += 1

        # Freeze panes (C5 -> first 2 cols + header rows stay visible)
        ws.freeze_panes = "C5"

        # Column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 50
        for j in range(n_years):
            ws.column_dimensions[get_column_letter(3 + j)].width = 13

        # Footer line
        footer_row = row + 1
        ws.cell(
            row=footer_row, column=1,
            value="LOW-priority rows in grey  -  Subtotals (codes with '/') in bold  -  '?m WARN' = non-standard period",
        ).font = FOOTER_FONT
        if last_col > 1:
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=last_col)

    def _build_filings_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Filings")
        d = self._data
        assert d is not None  # noqa: S101

        ws.cell(row=1, column=1, value="Filings (selected)").font = TITLE_FONT
        ws.merge_cells("A1:J1")

        headers = [
            "Year", "Period", "Period Label", "Model", "Reference",
            "Period Start", "Period End", "Deposit Date", "Source", "Enterprise Name",
        ]
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=j, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for i, f in enumerate(d.filings, start=4):
            ws.cell(row=i, column=1,  value=f.derived_year)
            ws.cell(row=i, column=2,  value=self._period_display(f.period_months))
            ws.cell(row=i, column=3,  value=f.period_label)
            ws.cell(row=i, column=4,  value=f.nbb_model_type or "-")
            ws.cell(row=i, column=5,  value=f.filing_reference)
            ws.cell(row=i, column=6,  value=f.period_start)
            ws.cell(row=i, column=7,  value=f.period_end)
            ws.cell(row=i, column=8,  value=f.deposit_date)
            ws.cell(row=i, column=9,  value=f.source_code)
            ws.cell(row=i, column=10, value=f.enterprise_name or "-")

        widths = [8, 12, 14, 10, 18, 12, 12, 14, 10, 40]
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w

        ws.freeze_panes = "A4"

    # ----------------------------------------------------- row-level helpers

    def _write_section_header(
        self, ws: Worksheet, row: int, last_col: int, section: str
    ) -> None:
        label = SECTION_HEADERS.get(section, section)
        ws.cell(row=row, column=1, value=label).font = SECTION_FONT
        for c in range(1, last_col + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        if last_col > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)

    def _write_code_row(
        self,
        ws: Worksheet,
        row: int,
        code: _CodeRow,
        filings: list[_Filing],
        amounts: dict[str, dict[str, float]],
    ) -> None:
        is_subtotal = "/" in code.pcmn_code
        font = (
            SUBTOTAL_FONT if is_subtotal
            else HIGH_FONT if code.v4g_priority == "HIGH"
            else LOW_FONT if code.v4g_priority == "LOW"
            else MED_FONT
        )
        fill = SUBTOTAL_FILL if is_subtotal else None

        ws.cell(row=row, column=1, value=code.pcmn_code).font = font
        ws.cell(row=row, column=2, value=code.label_en).font = font
        if fill is not None:
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=2).fill = fill

        # Rendering rule based on the code's subsection (from MAR_Dictionary):
        #   * subsection == 'Count'     -> raw value, integer format (e.g. 24,800 hours)
        #   * subsection == 'Headcount' -> raw value, decimal format (e.g. 8.5 FTE)
        #   * other                     -> divide by 1000, decimal format (EUR k)
        skip_scaling = code.subsection in COUNT_SUBSECTIONS
        use_integer = code.subsection == "Count"

        code_amounts = amounts.get(code.pcmn_code, {})
        for j, f in enumerate(filings):
            cell = ws.cell(row=row, column=3 + j)
            amt = code_amounts.get(f.filing_id)
            if amt is not None:
                cell.value = amt if skip_scaling else amt / 1000.0
                cell.number_format = COUNT_FORMAT if use_integer else AMOUNT_FORMAT
            # else: leave blank (deliberate -- NULL means N/A, not 0)
            cell.font = font
            cell.alignment = Alignment(horizontal="right")
            if fill is not None:
                cell.fill = fill

    @staticmethod
    def _period_flag(months: int | None) -> str:
        """Compact non-standard period flag for the year column header."""
        if months is None or months == 12:
            return ""
        return f"{months}m WARN"

    @staticmethod
    def _period_display(months: int | None) -> str:
        """Verbose period display for the Filings sheet."""
        if months is None:
            return "-"
        if months == 12:
            return "12m"
        return f"{months}m (non-standard)"
