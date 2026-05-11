"""V4G Excel export — Lane "Export" of /tools.

Reads canonical financials from Supabase (no NBB API call) and produces a
single-file .xlsx. Two output modes:

- ``simple``  — Info, P&L, Balance Sheet, KPIs (same surface as legacy
  v4g_accounts output; sufficient for client-facing decks).
- ``analyst`` — Simple sheets plus a Filings metadata sheet and one
  Detail_<year> sheet per most-recent filing with all PCMN line items
  resolved against ``dim_pcmn_codes``.

Source-of-truth tables:
  - ``party_registry``           — display_name lookup
  - ``party_identifiers``        — KBO → party_id resolution
  - ``fact_financials_evidence`` — KPI aggregates per period (EUR M)
  - ``fact_filings``             — per-filing metadata
  - ``fact_financials_lines``    — per-PCMN-code amounts per filing
  - ``dim_pcmn_codes``           — code → NL description, section

Typical usage::

    from src.services.excel_export import ExcelExporter
    from src.persistence.supabase import admin_client

    exporter = ExcelExporter(
        client=admin_client(),
        party_id="38cff812-397f-5fb4-bf18-a0e8b42b2a69",
        mode="analyst",
        year_limit=10,
    )
    exporter.fetch()
    content_bytes = exporter.build()

Failure modes:
  - ``ExportError`` if party_id has no evidence rows (nothing to export).
  - Underlying Supabase exceptions are propagated unchanged.

No I/O side effects: returns bytes; caller decides how to deliver.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

SOURCE_CODE_NBB = "SRC_NBB"
ExportMode = Literal["simple", "analyst"]


class ExportError(Exception):
    """No exportable data for the given party_id."""


# ─────────────────────────────────────────────────────────────────────────────
# Style tokens — flat, neutral, print-friendly
# ─────────────────────────────────────────────────────────────────────────────

_FONT_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FILL_HDR = PatternFill("solid", fgColor="2C2C2A")  # near-black (matches doc system)
_FILL_ALT = PatternFill("solid", fgColor="F1EFE8")  # very light neutral
_FONT_LBL = Font(name="Calibri", size=10, bold=True, color="2C2C2A")
_FONT_BDY = Font(name="Calibri", size=10)
_ALIGN_R = Alignment(horizontal="right", vertical="center")
_ALIGN_L = Alignment(horizontal="left", vertical="center")
_ALIGN_C = Alignment(horizontal="center", vertical="center")

_NUM_FMT_EUR_M = "#,##0.000;[Red]-#,##0.000"
_NUM_FMT_INT = "#,##0;[Red]-#,##0"
_NUM_FMT_PCT = "0.0%;[Red]-0.0%"
_NUM_FMT_DATE = "yyyy-mm-dd"


# ─────────────────────────────────────────────────────────────────────────────
# In-memory shapes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class _PartyMeta:
    party_id: str
    display_name: str
    kbo: str | None
    country: str = "BE"


@dataclass
class _EvidenceRow:
    period_label: str          # "2024"
    period_end: date | None
    revenue_eur_m: float | None
    ebitda_eur_m: float | None
    ebit_eur_m: float | None
    net_income_eur_m: float | None
    total_assets_eur_m: float | None
    total_equity_eur_m: float | None
    cash_eur_m: float | None
    total_debt_eur_m: float | None
    net_debt_eur_m: float | None
    employees: int | None


@dataclass
class _FilingRow:
    filing_id: str
    filing_reference: str
    period_label: str
    period_start: date | None
    period_end: date | None
    period_months: int | None
    period_flag: str | None       # 'normal'/'extended'/'shortened'
    nbb_model_type: str | None
    deposit_date: date | None
    enterprise_name: str | None
    loaded_at: datetime | None


@dataclass
class _LineRow:
    pcmn_code: str
    amount_eur: float
    data_type: str
    amount_period: str            # 'current'/'previous'


@dataclass
class _PcmnCode:
    pcmn_code: str
    description_nl: str | None
    section: str | None


@dataclass
class _ExportData:
    party: _PartyMeta
    evidence: list[_EvidenceRow] = field(default_factory=list)
    filings: list[_FilingRow] = field(default_factory=list)
    lines_by_filing: dict[str, list[_LineRow]] = field(default_factory=dict)
    pcmn: dict[str, _PcmnCode] = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────────────────────
# Exporter
# ─────────────────────────────────────────────────────────────────────────────

class ExcelExporter:
    """DB → Excel exporter for one party.

    Two phases: ``fetch()`` populates internal data caches via Supabase reads,
    ``build()`` returns Excel bytes. Split is deliberate — fetch is the slow
    part (network), build is pure CPU.
    """

    def __init__(
        self,
        client: Any,
        party_id: str,
        mode: ExportMode = "simple",
        year_limit: int = 10,
        source_code: str = SOURCE_CODE_NBB,
    ) -> None:
        if mode not in ("simple", "analyst"):
            raise ValueError(f"mode must be 'simple' or 'analyst', got {mode!r}")
        if year_limit < 1:
            raise ValueError(f"year_limit must be >= 1, got {year_limit}")

        self._client = client
        self.party_id = party_id
        self.mode: ExportMode = mode
        self.year_limit = year_limit
        self.source_code = source_code

        self._data: _ExportData | None = None

    # ── fetch ────────────────────────────────────────────────────────────

    def fetch(self) -> ExcelExporter:
        """Populate internal caches. Raises ``ExportError`` if no evidence."""
        log.info(
            "excel_export.fetch · party=%s mode=%s year_limit=%d",
            self.party_id, self.mode, self.year_limit,
        )

        party = self._fetch_party()
        evidence = self._fetch_evidence()
        if not evidence:
            raise ExportError(
                f"No evidence rows for party_id={self.party_id} "
                f"(source_code={self.source_code}). Run a Fetch first."
            )

        filings = self._fetch_filings() if self.mode == "analyst" else []
        lines_by_filing: dict[str, list[_LineRow]] = {}
        pcmn: dict[str, _PcmnCode] = {}

        if self.mode == "analyst" and filings:
            recent_ids = [f.filing_id for f in filings[: self.year_limit]]
            lines_by_filing = self._fetch_lines_for_filings(recent_ids)
            needed_codes = {ln.pcmn_code for lns in lines_by_filing.values() for ln in lns}
            pcmn = self._fetch_pcmn_codes(needed_codes)

        self._data = _ExportData(
            party=party,
            evidence=evidence,
            filings=filings,
            lines_by_filing=lines_by_filing,
            pcmn=pcmn,
        )
        log.info(
            "excel_export.fetch done · evidence=%d filings=%d "
            "lines_filings=%d pcmn_codes=%d",
            len(evidence), len(filings), len(lines_by_filing), len(pcmn),
        )
        return self

    def _fetch_party(self) -> _PartyMeta:
        # party_registry: display_name, country, etc.
        reg = (
            self._client.table("party_registry")
            .select("party_id, display_name, country_iso2")
            .eq("party_id", self.party_id)
            .limit(1)
            .execute()
        )
        if not reg.data:
            raise ExportError(f"party_id {self.party_id} not in party_registry")
        row = reg.data[0]

        ids = (
            self._client.table("party_identifiers")
            .select("id_value")
            .eq("party_id", self.party_id)
            .eq("id_type", "KBO")
            .limit(1)
            .execute()
        )
        kbo = ids.data[0]["id_value"] if ids.data else None

        return _PartyMeta(
            party_id=self.party_id,
            display_name=row.get("display_name") or f"Party {self.party_id[:8]}",
            kbo=kbo,
            country=row.get("country_iso2") or "BE",
        )

    def _fetch_evidence(self) -> list[_EvidenceRow]:
        cols = (
            "period_label, period_end, revenue_eur_m, ebitda_eur_m, ebit_eur_m, "
            "net_income_eur_m, total_assets_eur_m, total_equity_eur_m, "
            "cash_eur_m, total_debt_eur_m, net_debt_eur_m, employees"
        )
        res = (
            self._client.table("fact_financials_evidence")
            .select(cols)
            .eq("party_id", self.party_id)
            .eq("source_code", self.source_code)
            .order("period_end", desc=True)
            .limit(self.year_limit)
            .execute()
        )
        return [_evidence_from_row(r) for r in (res.data or [])]

    def _fetch_filings(self) -> list[_FilingRow]:
        cols = (
            "filing_id, filing_reference, period_label, period_start, period_end, "
            "period_months, period_flag, nbb_model_type, deposit_date, "
            "enterprise_name, loaded_at"
        )
        res = (
            self._client.table("fact_filings")
            .select(cols)
            .eq("party_id", self.party_id)
            .eq("source_code", self.source_code)
            .order("period_end", desc=True)
            .limit(self.year_limit)
            .execute()
        )
        return [_filing_from_row(r) for r in (res.data or [])]

    def _fetch_lines_for_filings(
        self, filing_ids: list[str],
    ) -> dict[str, list[_LineRow]]:
        if not filing_ids:
            return {}
        # Supabase .in_() — fetch all lines in one query, group client-side.
        res = (
            self._client.table("fact_financials_lines")
            .select("filing_id, pcmn_code, amount_eur, data_type, amount_period")
            .in_("filing_id", filing_ids)
            .eq("amount_period", "current")
            .execute()
        )
        out: dict[str, list[_LineRow]] = {fid: [] for fid in filing_ids}
        for r in res.data or []:
            fid = r["filing_id"]
            if fid not in out:
                continue
            out[fid].append(_LineRow(
                pcmn_code=r["pcmn_code"],
                amount_eur=float(r["amount_eur"]),
                data_type=r.get("data_type") or "",
                amount_period=r.get("amount_period") or "current",
            ))
        return out

    def _fetch_pcmn_codes(self, codes: set[str]) -> dict[str, _PcmnCode]:
        if not codes:
            return {}
        res = (
            self._client.table("dim_pcmn_codes")
            .select("pcmn_code, description_nl, section")
            .in_("pcmn_code", list(codes))
            .execute()
        )
        return {
            r["pcmn_code"]: _PcmnCode(
                pcmn_code=r["pcmn_code"],
                description_nl=r.get("description_nl"),
                section=r.get("section"),
            )
            for r in (res.data or [])
        }

    # ── build ────────────────────────────────────────────────────────────

    def build(self) -> bytes:
        """Render to Excel bytes. Call ``fetch()`` first."""
        if self._data is None:
            raise RuntimeError("Call fetch() before build()")

        wb = Workbook()
        # Replace default sheet
        wb.remove(wb.active)

        _build_info_sheet(wb, self._data, self.mode, self.source_code)
        _build_pl_sheet(wb, self._data)
        _build_bs_sheet(wb, self._data)
        _build_kpi_sheet(wb, self._data)

        if self.mode == "analyst":
            _build_filings_sheet(wb, self._data)
            _build_detail_sheets(wb, self._data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── convenience ──────────────────────────────────────────────────────

    def suggest_filename(self) -> str:
        """Build a safe filename: ``V4G_<NAME>_<KBO>_<mode>.xlsx``."""
        if self._data is None:
            raise RuntimeError("Call fetch() before suggest_filename()")
        name = self._data.party.display_name or "party"
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in name)
        safe = safe.strip().replace(" ", "_")[:30]
        kbo = self._data.party.kbo or self.party_id[:8]
        return f"V4G_{safe}_{kbo}_{self.mode}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# Row builders — convert API rows to dataclasses, normalizing date/number types
# ─────────────────────────────────────────────────────────────────────────────

def _evidence_from_row(r: dict) -> _EvidenceRow:
    return _EvidenceRow(
        period_label=str(r.get("period_label") or ""),
        period_end=_parse_date(r.get("period_end")),
        revenue_eur_m=_parse_num(r.get("revenue_eur_m")),
        ebitda_eur_m=_parse_num(r.get("ebitda_eur_m")),
        ebit_eur_m=_parse_num(r.get("ebit_eur_m")),
        net_income_eur_m=_parse_num(r.get("net_income_eur_m")),
        total_assets_eur_m=_parse_num(r.get("total_assets_eur_m")),
        total_equity_eur_m=_parse_num(r.get("total_equity_eur_m")),
        cash_eur_m=_parse_num(r.get("cash_eur_m")),
        total_debt_eur_m=_parse_num(r.get("total_debt_eur_m")),
        net_debt_eur_m=_parse_num(r.get("net_debt_eur_m")),
        employees=int(r["employees"]) if r.get("employees") is not None else None,
    )


def _filing_from_row(r: dict) -> _FilingRow:
    return _FilingRow(
        filing_id=str(r["filing_id"]),
        filing_reference=str(r.get("filing_reference") or ""),
        period_label=str(r.get("period_label") or ""),
        period_start=_parse_date(r.get("period_start")),
        period_end=_parse_date(r.get("period_end")),
        period_months=int(r["period_months"]) if r.get("period_months") is not None else None,
        period_flag=r.get("period_flag"),
        nbb_model_type=r.get("nbb_model_type"),
        deposit_date=_parse_date(r.get("deposit_date")),
        enterprise_name=r.get("enterprise_name"),
        loaded_at=_parse_datetime(r.get("loaded_at")),
    )


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    # datetime is a subclass of date — if we get one, extract just the date
    # part (also strips any tzinfo, which would otherwise crash openpyxl).
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return date.fromisoformat(str(v)[:10])
    except (ValueError, TypeError):
        return None


def _parse_datetime(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        # Excel does not support timezone-aware datetimes — strip tzinfo
        return v.replace(tzinfo=None) if v.tzinfo else v
    try:
        dt = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return dt.replace(tzinfo=None) if dt.tzinfo else dt
    except (ValueError, TypeError):
        return None


def _parse_num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_info_sheet(wb: Workbook, d: _ExportData, mode: ExportMode, src: str) -> None:
    ws = wb.create_sheet("Info")
    rows = [
        ("V4G Financial Analysis", ""),
        ("", ""),
        ("Party",           d.party.display_name),
        ("KBO",             d.party.kbo or "—"),
        ("Country",         d.party.country),
        ("Source",          src),
        ("Years covered",   len(d.evidence)),
        ("Most recent",     d.evidence[0].period_label if d.evidence else "—"),
        ("Oldest",          d.evidence[-1].period_label if d.evidence else "—"),
        ("Mode",            mode),
        ("Generated",       datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    # Title row
    ws["A1"].value = rows[0][0]
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells("A1:B1")

    for i, (label, value) in enumerate(rows[2:], start=3):
        ws.cell(row=i, column=1, value=label).font = _FONT_LBL
        ws.cell(row=i, column=2, value=value).font = _FONT_BDY

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50


def _build_pl_sheet(wb: Workbook, d: _ExportData) -> None:
    """Profit & Loss summary — one column per year, rows = KPIs."""
    ws = wb.create_sheet("P&L")
    headers = ["Metric (€M)"] + [e.period_label for e in d.evidence]
    _write_headers(ws, headers)

    metrics = [
        ("Revenue",     "revenue_eur_m"),
        ("EBITDA",      "ebitda_eur_m"),
        ("EBIT",        "ebit_eur_m"),
        ("Net income",  "net_income_eur_m"),
    ]
    for row_idx, (label, attr) in enumerate(metrics, start=2):
        ws.cell(row=row_idx, column=1, value=label).font = _FONT_LBL
        for col_idx, e in enumerate(d.evidence, start=2):
            v = getattr(e, attr)
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.number_format = _NUM_FMT_EUR_M
            cell.alignment = _ALIGN_R
            cell.font = _FONT_BDY

    _autosize_cols(ws, headers, min_w=10, label_w=18)


def _build_bs_sheet(wb: Workbook, d: _ExportData) -> None:
    ws = wb.create_sheet("Balance Sheet")
    headers = ["Metric (€M)"] + [e.period_label for e in d.evidence]
    _write_headers(ws, headers)

    metrics = [
        ("Total assets",        "total_assets_eur_m"),
        ("Total equity",        "total_equity_eur_m"),
        ("Cash",                "cash_eur_m"),
        ("Total debt",          "total_debt_eur_m"),
        ("Net debt",            "net_debt_eur_m"),
    ]
    for row_idx, (label, attr) in enumerate(metrics, start=2):
        ws.cell(row=row_idx, column=1, value=label).font = _FONT_LBL
        for col_idx, e in enumerate(d.evidence, start=2):
            v = getattr(e, attr)
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.number_format = _NUM_FMT_EUR_M
            cell.alignment = _ALIGN_R
            cell.font = _FONT_BDY

    _autosize_cols(ws, headers, min_w=10, label_w=22)


def _build_kpi_sheet(wb: Workbook, d: _ExportData) -> None:
    """Derived ratios — margins, leverage, capital structure."""
    ws = wb.create_sheet("KPIs")
    headers = ["Ratio"] + [e.period_label for e in d.evidence]
    _write_headers(ws, headers)

    def _safe_div(num: float | None, den: float | None) -> float | None:
        if num is None or den is None or den == 0:
            return None
        return num / den

    rows = [
        ("EBITDA margin",        lambda e: _safe_div(e.ebitda_eur_m, e.revenue_eur_m), _NUM_FMT_PCT),
        ("EBIT margin",          lambda e: _safe_div(e.ebit_eur_m, e.revenue_eur_m),   _NUM_FMT_PCT),
        ("Net margin",           lambda e: _safe_div(e.net_income_eur_m, e.revenue_eur_m), _NUM_FMT_PCT),
        ("Equity / Assets",      lambda e: _safe_div(e.total_equity_eur_m, e.total_assets_eur_m), _NUM_FMT_PCT),
        ("Net debt / EBITDA",    lambda e: _safe_div(e.net_debt_eur_m, e.ebitda_eur_m), _NUM_FMT_EUR_M),
        ("Employees",            lambda e: e.employees, _NUM_FMT_INT),
    ]
    for row_idx, (label, calc, fmt) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=label).font = _FONT_LBL
        for col_idx, e in enumerate(d.evidence, start=2):
            v = calc(e)
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.number_format = fmt
            cell.alignment = _ALIGN_R
            cell.font = _FONT_BDY

    _autosize_cols(ws, headers, min_w=10, label_w=22)


def _build_filings_sheet(wb: Workbook, d: _ExportData) -> None:
    """Per-filing metadata — analyst-only."""
    ws = wb.create_sheet("Filings")
    headers = [
        "Reference", "Period", "Start", "End", "Months", "Flag",
        "Model", "Deposit date", "Enterprise name", "Loaded at",
    ]
    _write_headers(ws, headers)

    for row_idx, f in enumerate(d.filings, start=2):
        row = [
            f.filing_reference,
            f.period_label,
            f.period_start,
            f.period_end,
            f.period_months,
            f.period_flag,
            f.nbb_model_type,
            f.deposit_date,
            f.enterprise_name,
            f.loaded_at,
        ]
        for col_idx, v in enumerate(row, start=1):
            v = _excel_safe(v)
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = _FONT_BDY
            if isinstance(v, (date, datetime)):
                cell.number_format = _NUM_FMT_DATE
            elif isinstance(v, int):
                cell.number_format = _NUM_FMT_INT
                cell.alignment = _ALIGN_R

    _autosize_cols(ws, headers, min_w=10, label_w=18)


def _excel_safe(v: Any) -> Any:
    """Coerce a value to something openpyxl can write.

    Excel does NOT support timezone-aware datetimes. Some Supabase responses
    return tz-aware ``datetime`` objects even for ``date`` columns (postgrest-py
    auto-deserialization quirk). This belt-and-suspenders helper guarantees
    that no tz-aware datetime ever reaches a cell.
    """
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    return v


def _build_detail_sheets(wb: Workbook, d: _ExportData) -> None:
    """One sheet per filing — analyst-only.

    Sorted: section first (alphabetical, NULL last), then PCMN code.
    """
    for f in d.filings:
        lines = d.lines_by_filing.get(f.filing_id, [])
        if not lines:
            continue

        sheet_name = f"Detail_{f.period_label}"[:31]  # Excel sheet name limit
        # Disambiguate if needed
        suffix = 1
        base = sheet_name
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base[:28]}_{suffix}"
            suffix += 1
        ws = wb.create_sheet(sheet_name)

        headers = ["Section", "PCMN code", "Description (NL)", "Amount (EUR)", "Data type"]
        _write_headers(ws, headers)

        def sort_key(ln: _LineRow) -> tuple[str, str]:
            code = d.pcmn.get(ln.pcmn_code)
            section = (code.section if code and code.section else "z_unknown").lower()
            return (section, ln.pcmn_code)

        for row_idx, ln in enumerate(sorted(lines, key=sort_key), start=2):
            code = d.pcmn.get(ln.pcmn_code)
            section = code.section if code else None
            desc = code.description_nl if code else None

            ws.cell(row=row_idx, column=1, value=section).font = _FONT_BDY
            ws.cell(row=row_idx, column=2, value=ln.pcmn_code).font = _FONT_BDY
            ws.cell(row=row_idx, column=3, value=desc).font = _FONT_BDY
            amt_cell = ws.cell(row=row_idx, column=4, value=ln.amount_eur)
            amt_cell.number_format = "#,##0.00;[Red]-#,##0.00"
            amt_cell.alignment = _ALIGN_R
            amt_cell.font = _FONT_BDY
            ws.cell(row=row_idx, column=5, value=ln.data_type).font = _FONT_BDY

        _autosize_cols(ws, headers, min_w=12, label_w=42, col_widths={3: 42, 4: 18})


# ─────────────────────────────────────────────────────────────────────────────
# Worksheet helpers
# ─────────────────────────────────────────────────────────────────────────────

def _write_headers(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _FONT_HDR
        cell.fill = _FILL_HDR
        cell.alignment = _ALIGN_C
    ws.freeze_panes = "B2"


def _autosize_cols(
    ws: Worksheet,
    headers: list[str],
    *,
    min_w: int = 10,
    label_w: int = 18,
    col_widths: dict[int, int] | None = None,
) -> None:
    """Approximate column auto-size. Excel doesn't expose true width without
    rendering; we use header length + a sensible padding instead.
    """
    explicit = col_widths or {}
    for col_idx in range(1, len(headers) + 1):
        if col_idx in explicit:
            ws.column_dimensions[get_column_letter(col_idx)].width = explicit[col_idx]
            continue
        if col_idx == 1:
            ws.column_dimensions[get_column_letter(col_idx)].width = label_w
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(
                min_w, len(headers[col_idx - 1]) + 2,
            )
