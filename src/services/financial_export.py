"""Financial export service.

Reads `public.vw_target_financials` (the analytical view with YoY deltas
+ ratios pre-computed in DB). Builds the 3-sheet Excel workbook used
by both the CLI download and the future Web-α `/api/.../export.xlsx` route.

Doctrine: analytics live in DB views. Python only queries, dedupes, and
formats. No business logic in templates or routes.

Source-conflict rule: if both SRC_NBB and SRC_PB rows exist for the same
period_label, SRC_NBB wins (NBB is authoritative for BE entities). This
is enforced at the service level via _dedupe_prefer_nbb — the view itself
returns both rows, by design (analyst tooling may want to see the diff).
"""
from __future__ import annotations

import io
import logging
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.persistence.supabase import admin_client

log = logging.getLogger(__name__)


# ─── Layout config ────────────────────────────────────────────────────────
# (column in vw_target_financials, display label, openpyxl number format)
# Rows starting with __sep__ render as a visual separator band.
PIVOT_METRICS: list[tuple[str, str | None, str | None]] = [
    ("revenue_eur_m",                "Revenue (EUR M)",            "#,##0.00"),
    ("revenue_yoy_delta_eur_m",      "  YoY Δ revenue (EUR M)",    "+#,##0.00;-#,##0.00"),
    ("ebitda_eur_m",                 "EBITDA (EUR M)",             "#,##0.00"),
    ("ebitda_yoy_delta_eur_m",       "  YoY Δ EBITDA (EUR M)",     "+#,##0.00;-#,##0.00"),
    ("ebitda_margin_pct",            "  EBITDA margin (%)",        "0.0"),
    ("ebit_eur_m",                   "EBIT (EUR M)",               "#,##0.00"),
    ("net_income_eur_m",             "Net income (EUR M)",         "#,##0.00"),
    ("__sep1__",                     None,                          None),
    ("total_assets_eur_m",           "Total assets (EUR M)",       "#,##0.00"),
    ("total_equity_eur_m",           "Total equity (EUR M)",       "#,##0.00"),
    ("cash_eur_m",                   "Cash (EUR M)",               "#,##0.00"),
    ("total_debt_eur_m",             "Total debt (EUR M)",         "#,##0.00"),
    ("net_debt_eur_m",               "Net debt (EUR M)",           "#,##0.00"),
    ("net_leverage_x",               "  Net leverage (x)",         "0.0"),
    ("working_capital_eur_m",        "Working capital (EUR M)",    "#,##0.00"),
    ("__sep2__",                     None,                          None),
    ("employees",                    "Employees (FTE)",            "#,##0"),
    ("revenue_per_employee_eur_k",   "  Revenue / FTE (EUR K)",    "#,##0"),
    ("ebitda_per_employee_eur_k",    "  EBITDA / FTE (EUR K)",     "#,##0"),
]

PROVENANCE_COLS = [
    "period_end", "period_label", "period_type",
    "source_code", "confidence",
    "nbb_model_type", "nbb_filing_date",
]

# V4G amber + dark navy palette
_HEADER_FILL = PatternFill("solid", fgColor="E8A020")
_HEADER_FONT = Font(bold=True, color="0F1520")
_TITLE_FONT  = Font(bold=True, size=14, color="E8A020")
_MUTED_FONT  = Font(color="64748B")
_SEP_FILL    = PatternFill("solid", fgColor="1E2D45")


# ─── Data fetch ───────────────────────────────────────────────────────────
def get_financial_history(party_id: str | UUID) -> list[dict]:
    """Fetch financial history for one party from vw_target_financials.

    Conflict resolution: when both SRC_NBB and SRC_PB rows exist for the
    same period_label, SRC_NBB wins. See _dedupe_prefer_nbb.

    Returns rows ordered by period_end DESC (most recent first).
    Empty list if no financial data — this is *not* an error.
    """
    client = admin_client()
    resp = (
        client.table("vw_target_financials")
        .select("*")
        .eq("party_id", str(party_id))
        .order("period_end", desc=True)
        .execute()
    )
    rows = resp.data or []
    return _dedupe_prefer_nbb(rows)


def _dedupe_prefer_nbb(rows: list[dict]) -> list[dict]:
    """Group by (period_label, period_end), prefer SRC_NBB on conflict.

    Composite key — period_label alone is insufficient for Belgian companies
    with extended (verlengd) or shortened (verkort) boekjaren, where two
    filings can legitimately share a period_label but represent different
    period boundaries. Examples:
      - FY change June→December: NBB filing for 2018-06-30 (regular) and
        2018-12-31 (6-month transition) both labeled "2018".
      - Cross-source mismatch: SRC_NBB has period_end=2018-06-30 (true
        extended FY), SRC_PB has period_end=2018-12-31 (calendar-year
        approximation). Both must survive — they describe different facts.

    Pure function, idempotent, preserves input order for first-seen keys.
    Tested in tests/services/test_financial_export.py — keep pure.
    """
    seen: dict[tuple, dict] = {}
    for row in rows:
        # NULL period_end normalized to "" so two label-only rows still group
        # together (avoids hash-mismatch on different NULL representations).
        key = (
            row.get("period_label"),
            str(row.get("period_end") or ""),
        )
        existing = seen.get(key)
        if existing is None:
            seen[key] = row
            continue
        # Conflict: prefer NBB
        existing_src = existing.get("source_code")
        new_src = row.get("source_code")
        if new_src == "SRC_NBB" and existing_src != "SRC_NBB":
            seen[key] = row
    return list(seen.values())


# ─── Sheet builders ───────────────────────────────────────────────────────
def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, min_w: int = 10, max_w: int = 28) -> None:
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)


def _party_display(party_meta: dict | None) -> str:
    """Prefer display_name, fall back to legal_name. Never empty."""
    if not party_meta:
        return "(unknown party)"
    return (
        party_meta.get("display_name")
        or party_meta.get("legal_name")
        or "(unknown party)"
    )


def build_pivot_sheet(ws, rows: list[dict], party_meta: dict | None) -> None:
    """Metrics as rows, fiscal years as columns. Chronological L→R."""
    ws.title = "Pivot"

    name = _party_display(party_meta)
    kbo  = (party_meta or {}).get("kbo_nr") or ""
    ws["A1"] = f"Financial pivot · {name}"
    ws["A1"].font = _TITLE_FONT
    sub = f"KBO: {kbo} · " if kbo else ""
    ws["A2"] = f"{sub}Source: vw_target_financials  ·  rows: {len(rows)}"
    ws["A2"].font = _MUTED_FONT

    chrono = sorted(rows, key=lambda r: r.get("period_end") or "9999-99-99")
    years = [r.get("period_label") or (r.get("period_end") or "")[:4] for r in chrono]

    ws.cell(row=4, column=1, value="Metric")
    for i, year in enumerate(years, start=2):
        ws.cell(row=4, column=i, value=year)
    _style_header_row(ws, 4, len(years) + 1)

    row_idx = 5
    for col, label, fmt in PIVOT_METRICS:
        if col.startswith("__sep"):
            for j in range(1, len(years) + 2):
                ws.cell(row=row_idx, column=j).fill = _SEP_FILL
            row_idx += 1
            continue
        ws.cell(row=row_idx, column=1, value=label)
        for j, r in enumerate(chrono, start=2):
            v = r.get(col)
            cell = ws.cell(row=row_idx, column=j, value=v)
            if fmt and v is not None:
                cell.number_format = fmt
        row_idx += 1

    ws.freeze_panes = "B5"
    _autosize(ws, min_w=14, max_w=22)


def build_raw_sheet(ws, rows: list[dict]) -> None:
    """Flat one-row-per-year dump. All columns from vw_target_financials."""
    ws.title = "Raw"
    if not rows:
        ws["A1"] = "(no rows)"
        return

    cols = list(rows[0].keys())
    for i, col in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=col)
    _style_header_row(ws, 1, len(cols))

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))

    ws.freeze_panes = "B2"
    _autosize(ws, min_w=10, max_w=24)


def build_provenance_sheet(ws, rows: list[dict]) -> None:
    """Audit trail per fiscal year — filing_date / source / confidence."""
    ws.title = "Provenance"

    ws["A1"] = "Per-period provenance"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "filing_date / source / confidence per fiscal year"
    ws["A2"].font = _MUTED_FONT

    for i, col in enumerate(PROVENANCE_COLS, start=1):
        ws.cell(row=4, column=i, value=col)
    _style_header_row(ws, 4, len(PROVENANCE_COLS))

    chrono = sorted(rows, key=lambda r: r.get("period_end") or "0000-00-00", reverse=True)
    for r_idx, row in enumerate(chrono, start=5):
        for c_idx, col in enumerate(PROVENANCE_COLS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))

    ws.freeze_panes = "A5"
    _autosize(ws, min_w=14, max_w=22)


# ─── Public API ───────────────────────────────────────────────────────────
def build_workbook(rows: list[dict], party_meta: dict | None) -> Workbook:
    """Build the 3-sheet (Pivot / Raw / Provenance) workbook in-memory."""
    wb = Workbook()
    build_pivot_sheet(wb.active, rows, party_meta)
    build_raw_sheet(wb.create_sheet(), rows)
    build_provenance_sheet(wb.create_sheet(), rows)
    return wb


def build_xlsx_bytes(rows: list[dict], party_meta: dict | None) -> bytes:
    """Render the workbook to bytes — for HTTP response delivery.

    Used by the future Flask route `/api/party/<uuid>/export.xlsx`.
    """
    wb = build_workbook(rows, party_meta)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def suggest_filename(party_meta: dict | None) -> str:
    """Construct a safe filename: V4G_<name>_<KBO>_financials.xlsx.

    Pure function — tested in tests/services/test_financial_export.py.
    """
    if party_meta is None:
        return "financials.xlsx"
    name = _party_display(party_meta)
    safe_name = "".join(c if c.isalnum() else "_" for c in name).strip("_")[:40]
    safe_name = safe_name or "party"
    kbo = party_meta.get("kbo_nr") or ""
    if kbo:
        return f"V4G_{safe_name}_{kbo}_financials.xlsx"
    return f"V4G_{safe_name}_financials.xlsx"
