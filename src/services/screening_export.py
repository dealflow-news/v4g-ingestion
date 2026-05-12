"""M&A first-cut screening export.

ScreeningExporter produces a compact single-sheet workbook giving an analyst
an at-a-glance investment readout: identity, snapshot, trend, ratios with
benchmarks, flags, and coverage quality. Not exhaustive -- decision-ready.

Differs from AnalystExporter:
* Fetches from ``fact_financials_evidence`` (aggregated KPIs, EUR millions)
  rather than ``fact_financials_lines`` (~80 PCMN codes). Faster, no pivot.
* One sheet, not three.
* Includes a flags/coverage layer for interpretation -- not just numbers.
* Replaces the legacy ExcelExporter ``mode="simple"`` codepath (which is
  slated for removal once this is in production).

Architecture:
* ``_compute_ratios()`` -- pure Python ratio math from evidence rows
* ``_evaluate_flags()`` -- applies a list of FlagRule objects (config-backed
  in v1; migratable to a ``dim_screening_flags`` table later)
* ``_analyze_coverage()`` -- assesses data-source quality vs economic profile

v1 ratios (computable from evidence only):
* Solvability (equity / total_assets)        -- Gauss benchmark mu=0.35, sigma=0.15
* EBITDA margin (ebitda / revenue)            -- rule-based flag band
* Net Debt / EBITDA (net_debt / ebitda)       -- rule-based flag band
* Revenue 3yr CAGR                            -- rule-based flag
* EBITDA 3yr CAGR                             -- rule-based flag
* Working capital intensity (wc / revenue)    -- rule-based flag band

Note: ``current_ratio`` deferred to a future version once fact_financials_evidence
gains current_assets/current_liabilities columns. ``working_capital / revenue``
is the v1 substitute liquidity signal.
"""

from __future__ import annotations

import logging
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from math import isfinite
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)


# ============================================================================
# Styling
# ============================================================================
TITLE_FONT     = Font(name="Calibri", size=14, bold=True, color="1F4E79")
META_FONT      = Font(name="Calibri", size=9,  italic=True, color="888888")
BANNER_FONT    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LABEL_FONT     = Font(name="Calibri", size=10, bold=True, color="1F4E79")
VALUE_FONT     = Font(name="Calibri", size=10)
VALUE_BOLD     = Font(name="Calibri", size=10, bold=True)
COMMENT_FONT   = Font(name="Calibri", size=9,  italic=True, color="555555")
FLAG_RED_FONT    = Font(name="Calibri", size=10, bold=True, color="C00000")
FLAG_YELLOW_FONT = Font(name="Calibri", size=10, bold=True, color="B7472A")
FLAG_GREEN_FONT  = Font(name="Calibri", size=10, bold=True, color="2E7D32")
FLAG_INFO_FONT   = Font(name="Calibri", size=10, color="555555")

BANNER_FILL    = PatternFill("solid", fgColor="1F4E79")
ROW_ALT_FILL   = PatternFill("solid", fgColor="F7F9FC")

EUR_M_FORMAT   = '#,##0.00 "M€";(#,##0.00) "M€";"-"'
PCT_FORMAT     = '0.0%;(0.0%);"-"'
RATIO_FORMAT   = '#,##0.00;(#,##0.00);"-"'
COUNT_FORMAT   = '#,##0;(#,##0);"-"'

# ============================================================================
# Benchmarks (v1 -- only confirmed sources)
# Source: V4G_NBB_CBSO_Connector_v03b.xlsx (Deal Screening sheet)
# ============================================================================
@dataclass(frozen=True)
class GaussBenchmark:
    mu: float
    sigma: float
    source: str = "v03b"

BENCHMARKS: dict[str, GaussBenchmark] = {
    "solvability": GaussBenchmark(mu=0.35, sigma=0.15),
    # Other ratios: rule-based flag bands only in v1 (no schijnprecisie).
}


# ============================================================================
# Flag rules (config-backed; migratable to dim_screening_flags later)
# ============================================================================
@dataclass(frozen=True)
class FlagRule:
    code: str
    severity: str         # 'red' / 'yellow' / 'info'
    label: str
    description: str
    condition: Callable[[dict, dict, _Coverage], bool]


def _is_negative_equity(latest: dict, _ratios: dict, _cov: _Coverage) -> bool:
    eq = latest.get("total_equity_eur_m")
    return eq is not None and eq < 0


def _is_leverage_high(_latest: dict, ratios: dict, _cov: _Coverage) -> bool:
    nd_ebitda = ratios.get("net_debt_to_ebitda")
    return nd_ebitda is not None and nd_ebitda > 4.0


def _is_leverage_elevated(_latest: dict, ratios: dict, _cov: _Coverage) -> bool:
    nd_ebitda = ratios.get("net_debt_to_ebitda")
    return nd_ebitda is not None and 2.5 < nd_ebitda <= 4.0


def _is_ebitda_thin(_latest: dict, ratios: dict, _cov: _Coverage) -> bool:
    margin = ratios.get("ebitda_margin")
    return margin is not None and margin < 0.05


def _is_ebitda_negative(latest: dict, _ratios: dict, _cov: _Coverage) -> bool:
    eb = latest.get("ebitda_eur_m")
    return eb is not None and eb < 0


def _is_revenue_declining(_latest: dict, ratios: dict, _cov: _Coverage) -> bool:
    cagr = ratios.get("revenue_cagr_3y")
    return cagr is not None and cagr < -0.05


def _is_ebitda_declining(_latest: dict, ratios: dict, _cov: _Coverage) -> bool:
    cagr = ratios.get("ebitda_cagr_3y")
    return cagr is not None and cagr < -0.10


def _is_solvability_weak(_latest: dict, ratios: dict, _cov: _Coverage) -> bool:
    """Solvability more than 2 sigma below Belgian KMO benchmark."""
    s = ratios.get("solvability")
    if s is None:
        return False
    bm = BENCHMARKS["solvability"]
    z = (s - bm.mu) / bm.sigma
    return z < -2.0


def _is_revenue_missing(latest: dict, _ratios: dict, cov: _Coverage) -> bool:
    """Revenue is NULL in the latest year (common with verkort/micro filings)."""
    return latest.get("revenue_eur_m") is None and cov.latest_year is not None


def _is_multi_source(_latest: dict, _ratios: dict, cov: _Coverage) -> bool:
    return len(cov.sources_used) > 1


def _is_partial_coverage(_latest: dict, _ratios: dict, cov: _Coverage) -> bool:
    return cov.years_covered_count < cov.years_requested


FLAG_RULES: list[FlagRule] = [
    # Red flags -- material concerns
    FlagRule("NEG_EQUITY",       "red",    "Negative equity",
             "Total equity is below zero in the latest period.",
             _is_negative_equity),
    FlagRule("LEVERAGE_HIGH",    "red",    "Leverage stressed",
             "Net Debt / EBITDA exceeds 4.0x.",
             _is_leverage_high),
    FlagRule("EBITDA_NEGATIVE",  "red",    "EBITDA negative",
             "Latest-year EBITDA is below zero.",
             _is_ebitda_negative),
    FlagRule("SOLVABILITY_WEAK", "red",    "Solvability below benchmark",
             "Solvability is more than 2 sigma below the Belgian KMO median.",
             _is_solvability_weak),

    # Yellow flags -- watch items
    FlagRule("LEVERAGE_ELEVATED", "yellow", "Leverage elevated",
             "Net Debt / EBITDA between 2.5x and 4.0x.",
             _is_leverage_elevated),
    FlagRule("EBITDA_THIN",       "yellow", "EBITDA margin thin",
             "EBITDA margin below 5%.",
             _is_ebitda_thin),
    FlagRule("REVENUE_DECLINING", "yellow", "Revenue declining",
             "3-year revenue CAGR is below -5%.",
             _is_revenue_declining),
    FlagRule("EBITDA_DECLINING",  "yellow", "EBITDA declining",
             "3-year EBITDA CAGR is below -10%.",
             _is_ebitda_declining),

    # Info flags -- data quality / context
    FlagRule("REVENUE_MISSING",  "info",   "Revenue not reported",
             "Latest year revenue is unavailable in NBB feed (typical for "
             "verkort/micro filings -- requires manual input).",
             _is_revenue_missing),
    FlagRule("MULTI_SOURCE",     "info",   "Multi-source timeline",
             "Timeline stitched from multiple sources -- check consistency.",
             _is_multi_source),
    FlagRule("PARTIAL_COVERAGE", "info",   "Partial year coverage",
             "Not all requested years are available.",
             _is_partial_coverage),
]


# ============================================================================
# Data shapes
# ============================================================================
@dataclass
class _PartyMeta:
    party_id: UUID
    display_name: str
    legal_name: str | None
    country_iso2: str | None
    party_type: str | None
    primary_kbo: str | None


@dataclass
class _EvidenceRow:
    period_label: str
    period_end: str | None
    period_type: str | None
    revenue_eur_m: float | None
    ebitda_eur_m: float | None
    ebit_eur_m: float | None
    net_income_eur_m: float | None
    total_assets_eur_m: float | None
    total_equity_eur_m: float | None
    cash_eur_m: float | None
    total_debt_eur_m: float | None
    net_debt_eur_m: float | None
    working_capital_eur_m: float | None
    employees: int | None
    source_code: str
    year: int | None       # derived from period_end


@dataclass
class _Flag:
    code: str
    severity: str
    label: str
    description: str


@dataclass
class _Coverage:
    years_requested: int
    years_covered: list[int]
    years_covered_count: int
    years_missing: list[int]
    latest_year: int | None
    latest_period_label: str | None
    sources_used: list[str]
    missing_fields_latest: list[str]
    confidence: str           # 'HIGH' / 'MEDIUM' / 'LOW'


@dataclass
class _Data:
    party: _PartyMeta
    evidence: list[_EvidenceRow] = field(default_factory=list)  # year desc
    ratios_by_year: dict[int, dict[str, float | None]] = field(default_factory=dict)
    latest_ratios: dict[str, float | None] = field(default_factory=dict)
    flags: list[_Flag] = field(default_factory=list)
    coverage: _Coverage | None = None


class ScreeningExportError(Exception):
    """Raised when the screening export cannot be produced."""


# ============================================================================
# Pure helpers
# ============================================================================
def _safe_div(a: float | None, b: float | None) -> float | None:
    if a is None or b is None or b == 0:
        return None
    val = a / b
    return val if isfinite(val) else None


def _cagr(latest: float | None, earlier: float | None, years: int) -> float | None:
    """Compound annual growth rate. Handles sign changes by returning None."""
    if latest is None or earlier is None or years <= 0:
        return None
    if earlier <= 0 or latest <= 0:
        # CAGR is undefined when crossing zero; return None to flag
        return None
    return (latest / earlier) ** (1.0 / years) - 1.0


# ============================================================================
# Main class
# ============================================================================
class ScreeningExporter:
    """Compact single-sheet M&A first-cut screening export."""

    def __init__(self, client: Any, party_id: UUID | str, year_limit: int = 5) -> None:
        self.client = client
        self.party_id = UUID(str(party_id)) if not isinstance(party_id, UUID) else party_id
        self.year_limit = max(2, year_limit)  # need at least 2 for trend
        self._data: _Data | None = None

    # ----------------------------------------------------------- public API
    def fetch(self) -> ScreeningExporter:
        party = self._fetch_party()
        evidence = self._fetch_evidence()
        if not evidence:
            raise ScreeningExportError(
                f"No fact_financials_evidence rows for party_id={self.party_id}"
            )
        ratios_by_year = self._compute_ratios(evidence)
        latest_year = evidence[0].year
        latest_ratios = ratios_by_year.get(latest_year, {}) if latest_year else {}
        coverage = self._analyze_coverage(evidence)
        flags = self._evaluate_flags(evidence[0], latest_ratios, coverage)
        self._data = _Data(
            party=party, evidence=evidence,
            ratios_by_year=ratios_by_year, latest_ratios=latest_ratios,
            flags=flags, coverage=coverage,
        )
        log.info(
            "screening_export.fetch: party=%s evidence_rows=%d flags=%d coverage=%s",
            self.party_id, len(evidence), len(flags), coverage.confidence,
        )
        return self

    def build(self) -> bytes:
        if self._data is None:
            raise ScreeningExportError("Call fetch() before build().")
        wb = Workbook()
        default = wb.active
        if default is not None:
            wb.remove(default)
        self._build_screening_sheet(wb)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def suggest_filename(self) -> str:
        if self._data is None:
            return f"V4G_screening_{self.party_id}.xlsx"
        safe = "".join(
            c if c.isalnum() or c in "-_" else "_"
            for c in (self._data.party.display_name or "unknown")
        )[:60]
        kbo = self._data.party.primary_kbo or "NOKBO"
        return f"V4G_{safe}_{kbo}_screening.xlsx"

    # ----------------------------------------------------- data fetching
    def _fetch_party(self) -> _PartyMeta:
        resp = (
            self.client.table("party_registry")
            .select("party_id, display_name, legal_name, country_iso2, party_type")
            .eq("party_id", str(self.party_id))
            .limit(1)
            .execute()
        )
        if not resp.data:
            raise ScreeningExportError(
                f"Party {self.party_id} not found in party_registry"
            )
        row = resp.data[0]
        kbo_resp = (
            self.client.table("party_identifiers")
            .select("id_value")
            .eq("party_id", str(self.party_id))
            .eq("id_type", "KBO")
            .limit(1)
            .execute()
        )
        primary_kbo = kbo_resp.data[0]["id_value"] if kbo_resp.data else None
        display_name = (
            row.get("display_name") or row.get("legal_name") or str(self.party_id)
        )
        return _PartyMeta(
            party_id=self.party_id,
            display_name=display_name,
            legal_name=row.get("legal_name"),
            country_iso2=row.get("country_iso2"),
            party_type=row.get("party_type"),
            primary_kbo=primary_kbo,
        )

    def _fetch_evidence(self) -> list[_EvidenceRow]:
        """Fetch fact_financials_evidence, dedupe by period_end (latest write wins),
        keep year_limit + 2 most recent rows for CAGR calculation."""
        resp = (
            self.client.table("fact_financials_evidence")
            .select(
                "period_label, period_end, period_type, revenue_eur_m, ebitda_eur_m, "
                "ebit_eur_m, net_income_eur_m, total_assets_eur_m, total_equity_eur_m, "
                "cash_eur_m, total_debt_eur_m, net_debt_eur_m, working_capital_eur_m, "
                "employees, source_code, updated_at"
            )
            .eq("party_id", str(self.party_id))
            .order("period_end", desc=True)
            .order("updated_at", desc=True)
            .execute()
        )
        rows = resp.data or []
        seen_ends: set[str] = set()
        deduped: list[_EvidenceRow] = []
        for r in rows:
            end = r.get("period_end")
            if end is None or end in seen_ends:
                continue
            seen_ends.add(end)
            try:
                year = int(str(end)[:4])
            except (ValueError, TypeError):
                year = None
            deduped.append(_EvidenceRow(
                period_label=str(r.get("period_label") or ""),
                period_end=end,
                period_type=r.get("period_type"),
                revenue_eur_m=_to_float(r.get("revenue_eur_m")),
                ebitda_eur_m=_to_float(r.get("ebitda_eur_m")),
                ebit_eur_m=_to_float(r.get("ebit_eur_m")),
                net_income_eur_m=_to_float(r.get("net_income_eur_m")),
                total_assets_eur_m=_to_float(r.get("total_assets_eur_m")),
                total_equity_eur_m=_to_float(r.get("total_equity_eur_m")),
                cash_eur_m=_to_float(r.get("cash_eur_m")),
                total_debt_eur_m=_to_float(r.get("total_debt_eur_m")),
                net_debt_eur_m=_to_float(r.get("net_debt_eur_m")),
                working_capital_eur_m=_to_float(r.get("working_capital_eur_m")),
                employees=int(r["employees"]) if r.get("employees") is not None else None,
                source_code=str(r.get("source_code") or ""),
                year=year,
            ))
        # Keep year_limit + a couple extra for CAGR look-back
        return deduped[: self.year_limit + 2]

    # ----------------------------------------------------- computations
    def _compute_ratios(
        self, evidence: list[_EvidenceRow]
    ) -> dict[int, dict[str, float | None]]:
        """Compute per-year ratios. Returns {year: {ratio_name: value}}."""
        result: dict[int, dict[str, float | None]] = {}
        # Index by year for CAGR look-back
        by_year: dict[int, _EvidenceRow] = {e.year: e for e in evidence if e.year is not None}

        for ev in evidence:
            if ev.year is None:
                continue
            ratios: dict[str, float | None] = {
                "solvability":             _safe_div(ev.total_equity_eur_m, ev.total_assets_eur_m),
                "ebitda_margin":           _safe_div(ev.ebitda_eur_m,        ev.revenue_eur_m),
                "net_debt_to_ebitda":      _safe_div(ev.net_debt_eur_m,      ev.ebitda_eur_m),
                "working_capital_intensity": _safe_div(ev.working_capital_eur_m, ev.revenue_eur_m),
                "revenue_cagr_3y":         None,
                "ebitda_cagr_3y":          None,
            }
            # CAGR vs 3 years earlier (if available)
            earlier_year = ev.year - 3
            if earlier_year in by_year:
                earlier = by_year[earlier_year]
                ratios["revenue_cagr_3y"] = _cagr(ev.revenue_eur_m, earlier.revenue_eur_m, 3)
                ratios["ebitda_cagr_3y"]  = _cagr(ev.ebitda_eur_m,  earlier.ebitda_eur_m,  3)
            result[ev.year] = ratios
        return result

    def _analyze_coverage(self, evidence: list[_EvidenceRow]) -> _Coverage:
        years_covered = sorted({e.year for e in evidence if e.year is not None}, reverse=True)
        latest = years_covered[0] if years_covered else None
        # "requested" range = year_limit years ending at latest
        expected = list(range(latest - self.year_limit + 1, latest + 1)) if latest else []
        years_missing = sorted(set(expected) - set(years_covered))
        sources = sorted({e.source_code for e in evidence if e.source_code})
        # Missing fields in latest year
        missing_fields: list[str] = []
        if evidence:
            latest_ev = evidence[0]
            for field_name, value in [
                ("revenue", latest_ev.revenue_eur_m),
                ("ebitda",  latest_ev.ebitda_eur_m),
                ("equity",  latest_ev.total_equity_eur_m),
                ("net_debt", latest_ev.net_debt_eur_m),
                ("employees", latest_ev.employees),
            ]:
                if value is None:
                    missing_fields.append(field_name)
        # Confidence heuristic
        if not missing_fields and not years_missing:
            confidence = "HIGH"
        elif len(missing_fields) <= 1 and len(years_missing) <= 1:
            confidence = "MEDIUM"
        else:
            confidence = "LOW"
        return _Coverage(
            years_requested=self.year_limit,
            years_covered=years_covered,
            years_covered_count=len(years_covered),
            years_missing=years_missing,
            latest_year=latest,
            latest_period_label=evidence[0].period_label if evidence else None,
            sources_used=sources,
            missing_fields_latest=missing_fields,
            confidence=confidence,
        )

    def _evaluate_flags(
        self, latest_ev: _EvidenceRow, latest_ratios: dict[str, float | None],
        coverage: _Coverage,
    ) -> list[_Flag]:
        latest_dict = {
            "revenue_eur_m":     latest_ev.revenue_eur_m,
            "ebitda_eur_m":      latest_ev.ebitda_eur_m,
            "ebit_eur_m":        latest_ev.ebit_eur_m,
            "net_income_eur_m":  latest_ev.net_income_eur_m,
            "total_assets_eur_m": latest_ev.total_assets_eur_m,
            "total_equity_eur_m": latest_ev.total_equity_eur_m,
            "net_debt_eur_m":    latest_ev.net_debt_eur_m,
        }
        out: list[_Flag] = []
        for rule in FLAG_RULES:
            try:
                if rule.condition(latest_dict, latest_ratios, coverage):
                    out.append(_Flag(rule.code, rule.severity, rule.label, rule.description))
            except Exception as e:  # noqa: BLE001 -- defensive; flag eval should never crash export
                log.warning("flag rule %s raised: %s", rule.code, e)
        return out

    # ----------------------------------------------------- sheet building
    def _build_screening_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Screening")
        d = self._data
        assert d is not None  # noqa: S101 -- guarded by build()

        # Column widths
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16

        row = 1
        row = self._write_header(ws, row)
        row = self._write_identity(ws, row + 1)
        row = self._write_snapshot(ws, row + 1)
        row = self._write_trend(ws, row + 1)
        row = self._write_ratios(ws, row + 1)
        row = self._write_flags(ws, row + 1)
        row = self._write_coverage(ws, row + 1)

        ws.freeze_panes = "A2"

    def _write_header(self, ws: Worksheet, row: int) -> int:
        d = self._data
        ws.cell(row=row, column=1, value=f"V4G Screening Export - {d.party.display_name}").font = TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        meta = (
            f"KBO: {d.party.primary_kbo or '-'}  ·  "
            f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}  ·  "
            f"Years requested: {self.year_limit}"
        )
        ws.cell(row=row, column=1, value=meta).font = META_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        return row

    def _write_banner(self, ws: Worksheet, row: int, label: str) -> int:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = BANNER_FONT
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = BANNER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        return row

    def _write_identity(self, ws: Worksheet, row: int) -> int:
        d = self._data
        row = self._write_banner(ws, row, "IDENTITY")
        rows = [
            ("Legal name",      d.party.legal_name or d.party.display_name),
            ("KBO",             d.party.primary_kbo or "-"),
            ("Party type",      d.party.party_type or "-"),
            ("Country",         d.party.country_iso2 or "-"),
            ("Years covered",   self._coverage_summary()),
            ("Latest period",   d.coverage.latest_period_label or "-"),
            ("Sources",         ", ".join(d.coverage.sources_used) or "-"),
        ]
        for label, value in rows:
            row += 1
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=2, value=value).font = VALUE_FONT
        return row

    def _coverage_summary(self) -> str:
        cov = self._data.coverage
        if not cov.years_covered:
            return "-"
        if len(cov.years_covered) == 1:
            return f"{cov.years_covered[0]} ({cov.years_covered_count} of {cov.years_requested})"
        return (
            f"{min(cov.years_covered)} - {max(cov.years_covered)} "
            f"({cov.years_covered_count} of {cov.years_requested})"
        )

    def _write_snapshot(self, ws: Worksheet, row: int) -> int:
        d = self._data
        ev = d.evidence[0] if d.evidence else None
        period = ev.period_label if ev else "-"
        row = self._write_banner(ws, row, f"SNAPSHOT - {period}")
        if ev is None:
            return row
        snapshot_rows: list[tuple[str, Any, str]] = [
            ("Revenue",       ev.revenue_eur_m,      EUR_M_FORMAT),
            ("EBITDA",        ev.ebitda_eur_m,       EUR_M_FORMAT),
            ("EBIT",          ev.ebit_eur_m,         EUR_M_FORMAT),
            ("Net Income",    ev.net_income_eur_m,   EUR_M_FORMAT),
            ("Employees",     ev.employees,          COUNT_FORMAT),
            ("Total Assets",  ev.total_assets_eur_m, EUR_M_FORMAT),
            ("Equity",        ev.total_equity_eur_m, EUR_M_FORMAT),
            ("Cash",          ev.cash_eur_m,         EUR_M_FORMAT),
            ("Total Debt",    ev.total_debt_eur_m,   EUR_M_FORMAT),
            ("Net Debt",      ev.net_debt_eur_m,     EUR_M_FORMAT),
        ]
        for label, value, fmt in snapshot_rows:
            row += 1
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            c = ws.cell(row=row, column=2)
            if value is not None:
                c.value = value
                c.number_format = fmt
            c.font = VALUE_BOLD if label in ("Revenue", "EBITDA", "Equity", "Net Debt") else VALUE_FONT
            c.alignment = Alignment(horizontal="right")
        return row

    def _write_trend(self, ws: Worksheet, row: int) -> int:
        d = self._data
        evidence = d.evidence[: self.year_limit]
        row = self._write_banner(ws, row, f"TREND - {len(evidence)} years")
        if len(evidence) < 2:
            row += 1
            ws.cell(row=row, column=1, value="(insufficient years for trend)").font = COMMENT_FONT
            return row

        # Header
        row += 1
        ws.cell(row=row, column=1, value="Metric").font = LABEL_FONT
        for j, ev in enumerate(evidence):
            c = ws.cell(row=row, column=2 + j, value=ev.year)
            c.font = LABEL_FONT
            c.alignment = Alignment(horizontal="center")
        # CAGR column header (only if 3+ years)
        if len(evidence) >= 3:
            cagr_col = 2 + len(evidence)
            ws.cell(row=row, column=cagr_col, value="3y CAGR").font = LABEL_FONT

        # Rows: Revenue, EBITDA, EBITDA margin %, Employees, Net Debt
        metrics: list[tuple[str, Callable[[_EvidenceRow], float | int | None], str]] = [
            ("Revenue",        lambda e: e.revenue_eur_m,                          EUR_M_FORMAT),
            ("EBITDA",         lambda e: e.ebitda_eur_m,                           EUR_M_FORMAT),
            ("EBITDA margin",  lambda e: _safe_div(e.ebitda_eur_m, e.revenue_eur_m), PCT_FORMAT),
            ("Employees",      lambda e: e.employees,                              COUNT_FORMAT),
            ("Net Debt",       lambda e: e.net_debt_eur_m,                         EUR_M_FORMAT),
        ]
        for label, getter, fmt in metrics:
            row += 1
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            for j, ev in enumerate(evidence):
                v = getter(ev)
                c = ws.cell(row=row, column=2 + j)
                if v is not None:
                    c.value = v
                    c.number_format = fmt
                c.alignment = Alignment(horizontal="right")
                c.font = VALUE_FONT
            # CAGR for Revenue, EBITDA, Employees
            if len(evidence) >= 3 and label in ("Revenue", "EBITDA", "Employees"):
                latest_v = getter(evidence[0])
                earlier_v = getter(evidence[min(3, len(evidence) - 1)])
                cagr_col = 2 + len(evidence)
                cagr_cell = ws.cell(row=row, column=cagr_col)
                if isinstance(latest_v, (int, float)) and isinstance(earlier_v, (int, float)):
                    cagr_val = _cagr(float(latest_v), float(earlier_v), 3)
                    if cagr_val is not None:
                        cagr_cell.value = cagr_val
                        cagr_cell.number_format = PCT_FORMAT
                cagr_cell.alignment = Alignment(horizontal="right")
                cagr_cell.font = VALUE_BOLD
        return row

    def _write_ratios(self, ws: Worksheet, row: int) -> int:
        d = self._data
        row = self._write_banner(ws, row, "RATIOS & BENCHMARKS")
        row += 1
        # Header row
        ws.cell(row=row, column=1, value="Ratio").font = LABEL_FONT
        ws.cell(row=row, column=2, value="Value").font = LABEL_FONT
        ws.cell(row=row, column=3, value="Benchmark").font = LABEL_FONT
        ws.cell(row=row, column=4, value="Z-score").font = LABEL_FONT
        ws.cell(row=row, column=5, value="Status").font = LABEL_FONT

        ratio_specs: list[tuple[str, str, str]] = [
            ("solvability",               "Solvability",              RATIO_FORMAT),
            ("ebitda_margin",             "EBITDA margin",            PCT_FORMAT),
            ("net_debt_to_ebitda",        "Net Debt / EBITDA",        RATIO_FORMAT),
            ("revenue_cagr_3y",           "Revenue 3y CAGR",          PCT_FORMAT),
            ("ebitda_cagr_3y",            "EBITDA 3y CAGR",           PCT_FORMAT),
            ("working_capital_intensity", "Working capital / Rev",    PCT_FORMAT),
        ]
        for key, label, fmt in ratio_specs:
            row += 1
            v = d.latest_ratios.get(key)
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            c = ws.cell(row=row, column=2)
            if v is not None:
                c.value = v
                c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
            c.font = VALUE_FONT

            bm = BENCHMARKS.get(key)
            if bm is not None:
                ws.cell(row=row, column=3,
                        value=f"μ={bm.mu:.2f}, σ={bm.sigma:.2f}").font = COMMENT_FONT
                if v is not None:
                    z = (v - bm.mu) / bm.sigma
                    zc = ws.cell(row=row, column=4, value=z)
                    zc.number_format = '+0.00"σ";-0.00"σ"'
                    zc.alignment = Alignment(horizontal="right")
                    zc.font = VALUE_FONT
                    status, fnt = self._gauss_status(z)
                    sc = ws.cell(row=row, column=5, value=status)
                    sc.font = fnt
            else:
                ws.cell(row=row, column=3, value="(rule-based)").font = COMMENT_FONT
        return row

    @staticmethod
    def _gauss_status(z: float) -> tuple[str, Font]:
        if z < -2.0:
            return "⚠ Very weak", FLAG_RED_FONT
        if z < -1.0:
            return "Below band", FLAG_YELLOW_FONT
        if z > 2.0:
            return "★ Exceptional", FLAG_GREEN_FONT
        if z > 1.0:
            return "Above band", FLAG_GREEN_FONT
        return "● Within band", FLAG_GREEN_FONT

    def _write_flags(self, ws: Worksheet, row: int) -> int:
        d = self._data
        row = self._write_banner(ws, row, "FLAGS")
        if not d.flags:
            row += 1
            ws.cell(row=row, column=1, value="● No critical flags raised").font = FLAG_GREEN_FONT
            return row
        # Sort: red first, then yellow, then info
        severity_order = {"red": 0, "yellow": 1, "info": 2}
        sorted_flags = sorted(d.flags, key=lambda f: severity_order.get(f.severity, 9))
        for flag in sorted_flags:
            row += 1
            icon, fnt = self._flag_icon(flag.severity)
            ws.cell(row=row, column=1, value=f"{icon} {flag.code}").font = fnt
            ws.cell(row=row, column=2, value=flag.label).font = fnt
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
            ws.cell(row=row, column=3, value=flag.description).font = COMMENT_FONT
        return row

    @staticmethod
    def _flag_icon(severity: str) -> tuple[str, Font]:
        if severity == "red":
            return "🔴", FLAG_RED_FONT
        if severity == "yellow":
            return "🟡", FLAG_YELLOW_FONT
        return "ℹ", FLAG_INFO_FONT

    def _write_coverage(self, ws: Worksheet, row: int) -> int:
        d = self._data
        cov = d.coverage
        row = self._write_banner(ws, row, "COVERAGE & DATA QUALITY")
        coverage_rows = [
            ("Years covered",
             f"{cov.years_covered_count} of {cov.years_requested} requested "
             f"({', '.join(str(y) for y in cov.years_covered)})" if cov.years_covered else "-"),
            ("Years missing",
             ", ".join(str(y) for y in cov.years_missing) if cov.years_missing else "-"),
            ("Missing fields (latest)",
             ", ".join(cov.missing_fields_latest) if cov.missing_fields_latest else "-"),
            ("Sources",
             ", ".join(cov.sources_used) if cov.sources_used else "-"),
            ("Confidence",
             cov.confidence),
        ]
        for label, value in coverage_rows:
            row += 1
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            v_cell = ws.cell(row=row, column=2, value=value)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            if label == "Confidence":
                if value == "HIGH":
                    v_cell.font = FLAG_GREEN_FONT
                elif value == "MEDIUM":
                    v_cell.font = FLAG_YELLOW_FONT
                else:
                    v_cell.font = FLAG_RED_FONT
            else:
                v_cell.font = VALUE_FONT
        return row


# ============================================================================
# Module-level helpers
# ============================================================================
def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None
